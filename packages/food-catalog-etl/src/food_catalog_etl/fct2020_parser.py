"""FCT2020 Excel のパースと FoodItem への変換。"""

from __future__ import annotations

import logging
from typing import Any

from fitness_contracts.models.food_item import FoodItem
from fitness_contracts.models.nutrient import NutrientQuality, NutrientValue

logger = logging.getLogger(__name__)

# FCT2020 列名 -> FoodItem フィールド名
COLUMN_MAP: dict[str, str] = {
    "エネルギー (kcal)": "energy_kcal",
    "たんぱく質": "protein_g",
    "脂質": "fat_g",
    "炭水化物": "carbs_g",
    "食物繊維総量": "fiber_g",
    "ナトリウム": "sodium_mg",
}


def parse_nutrient(raw: Any) -> NutrientValue:
    """セル値を NutrientValue に変換する。"""
    if raw is None:
        return NutrientValue(value=0.0, quality=NutrientQuality.MISSING)

    s = str(raw).strip()

    if s == "-":
        return NutrientValue(value=0.0, quality=NutrientQuality.MISSING)

    if s in ("Tr", "(Tr)"):
        return NutrientValue(value=0.0, quality=NutrientQuality.TRACE)

    # (0) は推定値ゼロ
    cleaned = s.strip("()")

    try:
        return NutrientValue(value=float(cleaned), quality=NutrientQuality.EXACT)
    except ValueError:
        return NutrientValue(value=0.0, quality=NutrientQuality.MISSING)


def parse_row(row: dict[str, Any], *, row_number: int) -> FoodItem | None:
    """行 dict を FoodItem に変換する。必須フィールド欠損時は None を返す。"""
    food_id_raw = row.get("食品番号")
    name_ja_raw = row.get("食品名")

    food_id = str(food_id_raw).strip() if food_id_raw is not None else ""
    name_ja = str(name_ja_raw).strip() if name_ja_raw is not None else ""

    if not food_id or not name_ja:
        logger.warning("Row %d: missing food_id or name_ja, skipping", row_number)
        return None

    nutrients = {
        field: parse_nutrient(row.get(col))
        for col, field in COLUMN_MAP.items()
    }

    try:
        return FoodItem(
            food_id=food_id,
            name_ja=name_ja,
            category=str(row.get("食品群", "")).strip(),
            **nutrients,
            source_row_number=row_number,
        )
    except Exception:
        logger.warning(
            "Row %d: validation failed, skipping", row_number, exc_info=True
        )
        return None


def parse_workbook(file_path: str) -> tuple[list[FoodItem], list[int]]:
    """Excel ファイルを読み込み、FoodItem リストとスキップ行番号リストを返す。"""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=False))
    if not rows:
        wb.close()
        return [], []

    # ヘッダー行を特定
    header_row = rows[0]
    headers = [cell.value for cell in header_row]

    # 必須列の存在チェック (フォーマット変更の最前線検知)
    required_columns = {"食品番号", "食品名", "食品群", *COLUMN_MAP.keys()}
    header_set = {h for h in headers if h is not None}
    missing_columns = required_columns - header_set
    if missing_columns:
        raise ValueError(
            f"FCT2020 Excel に必須列がありません: {sorted(missing_columns)}. "
            "ファイルのフォーマットが変更された可能性があります。"
        )

    items: list[FoodItem] = []
    skipped: list[int] = []

    for idx, row in enumerate(rows[1:], start=2):
        row_dict = dict(zip(headers, [cell.value for cell in row]))
        item = parse_row(row_dict, row_number=idx)
        if item is not None:
            items.append(item)
        else:
            skipped.append(idx)

    wb.close()
    return items, skipped
